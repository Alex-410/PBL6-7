#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = openpyxl.Workbook()

# 定义样式
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=12, color="FFFFFF")
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============ Sheet 1: 通过的测试用例 ============
ws_pass = wb.active
ws_pass.title = "通过的测试用例"

# 表头
pass_headers = ['编号', '模块', '功能名称', '测试用例名称', '级别', '测试步骤', '预期结果', '实际结果', '测试结果']
for col, header in enumerate(pass_headers, 1):
    cell = ws_pass.cell(row=1, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 通过的测试用例数据
pass_cases = [
    # 认证模块
    ['TC-AUTH-001', '认证模块', '用户注册', '正常注册新用户', '5',
     '1. 访问注册页面\n2. 输入用户名、密码、确认密码、邮箱、手机号\n3. 点击注册按钮',
     '注册成功，返回用户信息和token', '注册成功，返回code=200及token', 'Pass'],

    ['TC-AUTH-002', '认证模块', '用户注册', '重复用户名注册', '4',
     '1. 使用已存在的用户名"admin"进行注册\n2. 点击注册按钮',
     '提示"用户名已存在"，注册失败', '返回code=500，提示"用户名已存在"', 'Pass'],

    ['TC-AUTH-003', '认证模块', '用户注册', '密码不一致注册', '4',
     '1. 输入密码"Test123456"\n2. 输入确认密码"Different"\n3. 点击注册',
     '提示"两次输入的密码不一致"', '返回code=500，提示"两次输入的密码不一致"', 'Pass'],

    ['TC-AUTH-004', '认证模块', '用户注册', '弱密码注册', '3',
     '1. 输入密码"123"（少于6位）\n2. 点击注册',
     '提示密码长度不足，注册失败', '返回code=500，密码验证失败', 'Pass'],

    ['TC-AUTH-005', '认证模块', '用户注册', '无效邮箱注册', '3',
     '1. 输入邮箱"not-an-email"\n2. 点击注册',
     '提示邮箱格式不正确', '返回code=500，邮箱验证失败', 'Pass'],

    ['TC-AUTH-006', '认证模块', '用户注册', '无效手机号注册', '3',
     '1. 输入手机号"123"\n2. 点击注册',
     '提示手机号格式不正确', '返回code=500，手机号验证失败', 'Pass'],

    ['TC-AUTH-007', '认证模块', '用户登录', '不存在用户登录', '4',
     '1. 输入不存在的用户名"nonexist"\n2. 输入任意密码\n3. 点击登录',
     '提示"用户名或学号不存在"', '返回code=500，提示"用户名或学号不存在"', 'Pass'],

    ['TC-AUTH-008', '认证模块', '用户登录', '错误密码登录', '4',
     '1. 输入正确用户名\n2. 输入错误密码\n3. 点击登录',
     '提示"用户名或密码错误"', '返回code=500，提示"用户名或密码错误"', 'Pass'],

    ['TC-AUTH-009', '认证模块', '用户登录', '正确登录', '5',
     '1. 输入正确用户名\n2. 输入正确密码\n3. 点击登录',
     '登录成功，返回用户信息和token', '登录成功，返回code=200及token', 'Pass'],

    # 活动模块
    ['TC-ACT-001', '活动模块', '活动列表', '未认证访问活动列表', '4',
     '1. 不携带Authorization头\n2. GET请求/api/activities',
     '返回401或403，拒绝访问', '返回401 Unauthorized', 'Pass'],

    ['TC-ACT-002', '活动模块', '活动列表', '认证后访问活动列表', '5',
     '1. 携带有效token\n2. GET请求/api/activities',
     '返回200，包含活动列表数据', '返回code=200，包含20条活动数据', 'Pass'],

    ['TC-ACT-003', '活动模块', '创建活动', '正常创建活动', '5',
     '1. 携带有效token\n2. POST请求/api/activities\n3. 包含完整活动信息',
     '创建成功，返回活动详情', '返回code=200，返回活动ID和状态', 'Pass'],

    ['TC-ACT-004', '活动模块', '创建活动', '缺少标题创建活动', '4',
     '1. 携带有效token\n2. POST请求/api/activities\n3. 不包含title字段',
     '返回错误，提示标题不能为空', '返回code=500，验证失败', 'Pass'],

    ['TC-ACT-005', '活动模块', '活动详情', '获取活动详情', '4',
     '1. 携带有效token\n2. GET请求/api/activities/{id}',
     '返回200，包含活动详细信息', '返回code=200，包含活动详情', 'Pass'],

    # 报名模块
    ['TC-REG-001', '报名模块', '我的报名', '查看我的报名列表', '4',
     '1. 携带有效token\n2. GET请求/api/registrations/me',
     '返回200，包含当前用户的报名记录', '返回code=200，包含报名列表', 'Pass'],

    ['TC-REG-002', '报名模块', '活动报名', '报名不存在的活动', '4',
     '1. 携带有效token\n2. POST请求/api/registrations?activityId=99999',
     '返回错误，提示活动不存在', '返回code=500，提示活动不存在', 'Pass'],

    # 用户管理
    ['TC-USER-001', '用户管理', '用户详情', '获取不存在用户', '3',
     '1. 携带有效token\n2. GET请求/api/users/99999',
     '返回错误，提示用户不存在', '返回code=500，提示用户不存在', 'Pass'],

    ['TC-USER-002', '用户管理', '用户状态', '无效状态类型更新', '3',
     '1. 携带管理员token\n2. PUT请求/api/users/{id}/status\n3. body: {"status": "invalid"}',
     '返回友好错误提示', '返回错误提示，非500崩溃', 'Pass'],

    # RAG系统
    ['TC-RAG-001', 'RAG系统', '服务状态', 'RAG服务状态检查', '5',
     '1. GET请求/api/rag/status/',
     '返回200，包含索引状态和文档数', '返回has_index和doc_count', 'Pass'],

    ['TC-RAG-002', 'RAG系统', '对话功能', '空问题发送', '4',
     '1. POST请求/api/rag/chat/\n2. body: {"question": ""}',
     '返回400，提示请输入问题', '返回400 Bad Request', 'Pass'],

    ['TC-RAG-003', 'RAG系统', '对话管理', '创建新对话', '4',
     '1. POST请求/api/rag/conversations/\n2. body: {"title": "测试"}',
     '返回201，包含对话ID', '返回201 Created，包含对话ID', 'Pass'],

    ['TC-RAG-004', 'RAG系统', '文档管理', '删除不存在文档', '3',
     '1. POST请求/api/rag/delete/\n2. body: {"filename": "nonexist.txt"}',
     '返回400，提示文件不存在', '返回400，提示文件不存在', 'Pass'],

    # 安全测试
    ['TC-SEC-001', '安全测试', 'SQL注入', 'SQL注入防护', '5',
     '1. POST请求/api/auth/login\n2. username: "admin\' OR \'1\'=\'1"',
     '登录失败，SQL注入无效', '返回code=500，登录失败', 'Pass'],

    ['TC-SEC-002', '安全测试', 'JWT安全', 'JWT token篡改', '5',
     '1. 使用伪造token "fake.token.here"\n2. GET请求/api/activities',
     '返回401或403，拒绝访问', '返回401 Unauthorized', 'Pass'],
]

for row_idx, case in enumerate(pass_cases, 2):
    for col_idx, value in enumerate(case, 1):
        cell = ws_pass.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx == len(case):  # 测试结果列
            cell.fill = pass_fill
            cell.font = Font(bold=True, color="006100")

# 设置列宽
col_widths = [12, 10, 12, 20, 8, 40, 25, 25, 10]
for i, width in enumerate(col_widths, 1):
    ws_pass.column_dimensions[get_column_letter(i)].width = width

# ============ Sheet 2: 失败的测试用例（BUG） ============
ws_fail = wb.create_sheet("失败的测试用例(BUG)")

# 表头
fail_headers = ['编号', '模块', '功能名称', '测试用例名称', '级别', '测试步骤', '预期结果', '实际结果', '测试结果', 'BUG描述', '修复建议']
for col, header in enumerate(fail_headers, 1):
    cell = ws_fail.cell(row=1, column=col, value=header)
    cell.font = header_font_white
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 失败的测试用例数据（BUG）
fail_cases = [
    # 安全漏洞
    ['TC-FAIL-001', '认证模块', '登录安全', '登录频率限制测试', '5',
     '1. 连续10次使用错误密码登录同一用户\n2. 观察是否被限制',
     '超过5次后应锁定账户或IP，返回429',
     '10次全部返回500，无任何限制',
     'Fail',
     '无登录失败频率限制，攻击者可无限尝试暴力破解密码',
     '1. 添加登录失败计数器\n2. 超过5次失败后锁定15分钟\n3. 返回429 Too Many Requests'],

    ['TC-FAIL-002', '安全测试', 'JWT安全', '禁用用户后token有效性', '5',
     '1. 用户A登录获取token\n2. 管理员禁用用户A\n3. 用户A使用原token访问API',
     '返回401，token应立即失效',
     '返回200，token仍然有效，用户A仍可访问系统',
     'Fail',
     '禁用用户后JWT token仍然有效，缺少token失效机制',
     '1. 在JwtAuthenticationFilter中检查用户状态\n2. 或使用Redis维护token黑名单\n3. 禁用用户时清除其所有token'],

    ['TC-FAIL-003', '安全测试', 'XSS防护', '活动标题XSS注入', '4',
     '1. 创建活动，标题为"<script>alert(1)</script>"\n2. 查看返回数据',
     '标题应被转义为"&lt;script&gt;"',
     '标题原样返回"<script>alert(1)</script>"',
     'Fail',
     '活动标题未转义HTML标签，存在XSS攻击风险',
     '1. 入库前对用户输入进行HTML转义\n2. 输出时使用安全的渲染方式\n3. 前端使用v-text而非v-html'],

    # 权限控制缺陷
    ['TC-FAIL-004', '活动模块', '活动审核', '非管理员审核活动', '5',
     '1. 使用普通用户token\n2. PUT请求/api/activities/{id}/audit?action=approve',
     '返回403，提示无权限',
     '返回200，审核成功',
     'Fail',
     '任何用户都能审核活动，缺少管理员权限检查',
     '1. 在SecurityConfig中配置/admin/**需要ADMIN角色\n2. 在Controller中添加@PreAuthorize注解\n3. 或在Service层验证用户角色'],

    ['TC-FAIL-005', '活动模块', '活动删除', '删除他人活动', '5',
     '1. 使用用户A的token\n2. 删除用户B创建的活动',
     '返回403，提示无权删除',
     '返回200，删除成功',
     'Fail',
     '任何用户都能删除任意活动，缺少所有权检查',
     '1. 删除前验证当前用户是否为活动创建者\n2. 或验证是否为管理员\n3. 返回403 Forbidden'],

    ['TC-FAIL-006', '用户管理', '用户列表', '普通用户查看用户列表', '4',
     '1. 使用普通用户token\n2. GET请求/api/users',
     '返回403，提示无权限',
     '返回200，返回所有用户信息',
     'Fail',
     '普通用户可以查看所有用户列表，用户信息泄露风险',
     '1. 限制/users接口仅管理员可访问\n2. 使用@PreAuthorize("hasRole(\'ADMIN\')")注解'],

    # 逻辑缺陷
    ['TC-FAIL-007', '活动模块', '活动审核', '无效审核操作', '3',
     '1. 使用管理员token\n2. PUT请求/api/activities/{id}/audit?action=invalid',
     '返回400，提示无效操作',
     '返回200，默认执行为rejected',
     'Fail',
     'action参数无验证，任意值都会执行，应只允许approve/reject',
     '1. 在Controller中验证action值\n2. 非approve/reject返回400 Bad Request'],

    ['TC-FAIL-008', 'RAG系统', '文本分块', 'chunk_size等于overlap', '3',
     '1. 调用chunk_text函数\n2. 参数: chunk_size=100, overlap=100',
     '应报错或自动调整参数',
     '函数进入无限循环',
     'Fail',
     'fixed策略在chunk_size=overlap时start变量不前进，导致死循环',
     '1. 添加参数验证: overlap必须小于chunk_size\n2. 或在循环中添加安全退出条件'],
]

for row_idx, case in enumerate(fail_cases, 2):
    for col_idx, value in enumerate(case, 1):
        cell = ws_fail.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx == 9:  # 测试结果列
            cell.fill = fail_fill
            cell.font = Font(bold=True, color="9C0006")

# 设置列宽
fail_col_widths = [12, 10, 12, 20, 8, 40, 25, 25, 10, 35, 35]
for i, width in enumerate(fail_col_widths, 1):
    ws_fail.column_dimensions[get_column_letter(i)].width = width

# ============ Sheet 3: 测试汇总 ============
ws_summary = wb.create_sheet("测试汇总")

summary_headers = ['统计项', '数量', '占比']
for col, header in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=1, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

total = len(pass_cases) + len(fail_cases)
summary_data = [
    ['测试用例总数', total, '100%'],
    ['通过用例数', len(pass_cases), f'{len(pass_cases)/total*100:.1f}%'],
    ['失败用例数(BUG)', len(fail_cases), f'{len(fail_cases)/total*100:.1f}%'],
    ['', '', ''],
    ['BUG分布', '', ''],
    ['严重安全漏洞', 3, '37.5%'],
    ['权限控制缺陷', 3, '37.5%'],
    ['逻辑缺陷', 2, '25.0%'],
]

for row_idx, data in enumerate(summary_data, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 设置列宽
for i, width in enumerate([20, 15, 15], 1):
    ws_summary.column_dimensions[get_column_letter(i)].width = width

# 保存文件
output_file = '测试用例及跟踪_校园活动平台.xlsx'
wb.save(output_file)
print(f'测试用例文件已生成: {output_file}')
print(f'通过用例: {len(pass_cases)} 条')
print(f'失败用例: {len(fail_cases)} 条 (BUG)')
print(f'总计: {total} 条')

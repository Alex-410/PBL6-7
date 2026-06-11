#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import requests
import time
import json
import os
from dataclasses import dataclass, field
from datetime import datetime

# ============ 测试环境配置 ============
BASE_URL = 'http://localhost:8080/api'
RAG_URL = 'http://localhost:9001/api/rag'
TEST_TIME = datetime.now().strftime('%Y-%m-%d %H:%M:%S')


@dataclass
class TestCase:
    id: str
    module: str
    feature: str
    name: str
    priority: int
    steps: list = field(default_factory=list)
    expected: str = ''
    actual: str = ''
    result: str = ''
    request_detail: str = ''
    bug_desc: str = ''
    fix_suggestion: str = ''
    severity: str = ''


def run_test(tc: TestCase, condition: bool, actual: str = '', bug_desc: str = '', fix_suggestion: str = '', severity: str = ''):
    tc.actual = actual
    tc.result = 'Pass' if condition else 'Fail'
    if not condition:
        tc.bug_desc = bug_desc
        tc.fix_suggestion = fix_suggestion
        tc.severity = severity
    all_tests.append(tc)
    return condition


def print_test_result(tc: TestCase):
    status = '[PASS]' if tc.result == 'Pass' else '[FAIL]'
    print(f'{status} {tc.id}: {tc.module} > {tc.feature} > {tc.name}')
    if tc.request_detail:
        print(f'  请求: {tc.request_detail}')
    print(f'  测试账号: {tc.steps[0] if tc.steps else "N/A"}')
    print(f'  复现步骤:')
    for i, step in enumerate(tc.steps, 1):
        print(f'    {i}. {step}')
    print(f'  期望结果: {tc.expected}')
    print(f'  实际结果: {tc.actual}')
    if tc.result == 'Fail' and tc.bug_desc:
        print(f'  BUG描述: {tc.bug_desc}')
    print()


def print_bug_report():
    bugs = [tc for tc in all_tests if tc.result == 'Fail']
    if not bugs:
        print('未发现BUG')
        return

    print()
    print('=' * 70)
    print('                        BUG 报告')
    print('=' * 70)
    print(f'测试环境: {BASE_URL}')
    print(f'RAG服务:  {RAG_URL}')
    print(f'测试时间: {TEST_TIME}')
    print(f'测试账号: {test_username} / {test_password} (普通用户)')
    print('=' * 70)
    print()

    for i, tc in enumerate(bugs, 1):
        print(f'【BUG-{i:03d}】{tc.name}')
        print(f'  编号:     {tc.id}')
        print(f'  模块:     {tc.module} > {tc.feature}')
        print(f'  严重程度: {tc.severity}')
        print(f'  测试环境: {BASE_URL}')
        print(f'  测试账号: {test_username} / {test_password}')
        print(f'  复现步骤:')
        for j, step in enumerate(tc.steps, 1):
            print(f'    {j}. {step}')
        if tc.request_detail:
            print(f'  请求详情: {tc.request_detail}')
        print(f'  期望结果: {tc.expected}')
        print(f'  实际结果: {tc.actual}')
        print(f'  BUG描述: {tc.bug_desc}')
        print(f'  修复建议: {tc.fix_suggestion}')
        print()


# ============ 初始化 ============
all_tests = []
ts = str(int(time.time()))[-4:]
test_username = f'u{ts}'
test_password = 'Test123456'

# 打印测试环境信息
print('=' * 70)
print('               校园活动发布平台 - 完整测试报告')
print('=' * 70)
print(f'测试环境:   {BASE_URL}')
print(f'RAG服务:    {RAG_URL}')
print(f'测试时间:   {TEST_TIME}')
print('=' * 70)
print()

# ============ 注册测试账号 ============
reg_data = {'username': test_username, 'password': test_password,
            'confirmPassword': test_password,
            'email': f'{test_username}@test.com', 'phone': f'1390000{ts}'}
reg_resp = requests.post(f'{BASE_URL}/auth/register', json=reg_data)
reg_r = reg_resp.json()
user_token = reg_r.get('data', {}).get('token') if reg_r.get('data') else None
headers = {'Authorization': f'Bearer {user_token}'} if user_token else {}

print(f'测试账号:   {test_username} / {test_password} (普通用户)')
print(f'注册状态:   {"成功" if user_token else "失败"}')
print()

# ============ 认证模块 ============
print('=' * 70)
print('认证模块测试')
print('=' * 70)

# TC-001: 正常注册
tc = TestCase(
    id='TC-001', module='认证模块', feature='用户注册', name='正常注册新用户', priority=5,
    steps=[f'POST {BASE_URL}/auth/register', f'Body: {json.dumps(reg_data, ensure_ascii=False)}', '检查返回 code 是否为 200'],
    expected='code=200, 返回用户信息和token',
    request_detail=f'POST {BASE_URL}/auth/register'
)
run_test(tc, reg_r.get('code') == 200 and reg_r.get('data') is not None,
         actual=f'code={reg_r.get("code")}, token={"存在" if user_token else "无"}',
         bug_desc='注册接口异常', fix_suggestion='检查注册逻辑', severity='高')
print_test_result(tc)

# TC-002: 重复用户名
dup_data = {'username': test_username, 'password': 'Test123456', 'confirmPassword': 'Test123456',
            'email': 'other@test.com', 'phone': '13900009999'}
tc = TestCase(
    id='TC-002', module='认证模块', feature='用户注册', name='重复用户名注册', priority=4,
    steps=[f'POST {BASE_URL}/auth/register', f'Body: {json.dumps(dup_data, ensure_ascii=False)}', '检查返回 code 是否不为 200'],
    expected='code!=200, 提示用户名已存在',
    request_detail=f'POST {BASE_URL}/auth/register'
)
resp = requests.post(f'{BASE_URL}/auth/register', json=dup_data)
run_test(tc, resp.json().get('code') != 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='重复用户名未校验', fix_suggestion='注册前查询用户名是否已存在', severity='高')
print_test_result(tc)

# TC-003: 密码不一致
ts2 = str(int(time.time()) + 1)[-4:]
diff_data = {'username': f'u{ts2}', 'password': 'Test123456', 'confirmPassword': 'Different',
             'email': f'u{ts2}@test.com', 'phone': f'1390001{ts2}'}
tc = TestCase(
    id='TC-003', module='认证模块', feature='用户注册', name='密码不一致注册', priority=4,
    steps=[f'POST {BASE_URL}/auth/register', f'Body: {json.dumps(diff_data, ensure_ascii=False)}', '检查返回 code 是否不为 200'],
    expected='code!=200, 提示密码不一致',
    request_detail=f'POST {BASE_URL}/auth/register'
)
resp = requests.post(f'{BASE_URL}/auth/register', json=diff_data)
run_test(tc, resp.json().get('code') != 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='密码一致性未校验', fix_suggestion='注册时校验密码和确认密码是否一致', severity='中')
print_test_result(tc)

# TC-004: 弱密码
weak_data = {'username': f'w{ts2}', 'password': '123', 'confirmPassword': '123',
             'email': f'w{ts2}@test.com', 'phone': f'1390002{ts2}'}
tc = TestCase(
    id='TC-004', module='认证模块', feature='用户注册', name='弱密码注册', priority=3,
    steps=[f'POST {BASE_URL}/auth/register', f'Body: {json.dumps(weak_data, ensure_ascii=False)}', '检查返回 code 是否不为 200'],
    expected='code!=200, 提示密码长度不足',
    request_detail=f'POST {BASE_URL}/auth/register'
)
resp = requests.post(f'{BASE_URL}/auth/register', json=weak_data)
run_test(tc, resp.json().get('code') != 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='弱密码未校验', fix_suggestion='添加密码最小长度和复杂度验证', severity='中')
print_test_result(tc)

# TC-005: 无效邮箱
email_data = {'username': f'e{ts2}', 'password': 'Test123456', 'confirmPassword': 'Test123456',
              'email': 'not-an-email', 'phone': f'1390003{ts2}'}
tc = TestCase(
    id='TC-005', module='认证模块', feature='用户注册', name='无效邮箱注册', priority=3,
    steps=[f'POST {BASE_URL}/auth/register', f'Body: {json.dumps(email_data, ensure_ascii=False)}', '检查返回 code 是否不为 200'],
    expected='code!=200, 提示邮箱格式不正确',
    request_detail=f'POST {BASE_URL}/auth/register'
)
resp = requests.post(f'{BASE_URL}/auth/register', json=email_data)
run_test(tc, resp.json().get('code') != 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='邮箱格式未校验', fix_suggestion='使用正则表达式验证邮箱格式', severity='低')
print_test_result(tc)

# TC-006: 无效手机号
phone_data = {'username': f'p{ts2}', 'password': 'Test123456', 'confirmPassword': 'Test123456',
              'email': f'p{ts2}@test.com', 'phone': '123'}
tc = TestCase(
    id='TC-006', module='认证模块', feature='用户注册', name='无效手机号注册', priority=3,
    steps=[f'POST {BASE_URL}/auth/register', f'Body: {json.dumps(phone_data, ensure_ascii=False)}', '检查返回 code 是否不为 200'],
    expected='code!=200, 提示手机号格式不正确',
    request_detail=f'POST {BASE_URL}/auth/register'
)
resp = requests.post(f'{BASE_URL}/auth/register', json=phone_data)
run_test(tc, resp.json().get('code') != 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='手机号格式未校验', fix_suggestion='使用正则表达式验证手机号格式', severity='低')
print_test_result(tc)

# TC-007: 不存在用户登录
tc = TestCase(
    id='TC-007', module='认证模块', feature='用户登录', name='不存在用户登录', priority=4,
    steps=[f'POST {BASE_URL}/auth/login', 'Body: {"username":"nonexist","password":"test"}', '检查返回 code 是否不为 200'],
    expected='code!=200, 提示用户不存在',
    request_detail=f'POST {BASE_URL}/auth/login'
)
resp = requests.post(f'{BASE_URL}/auth/login', json={'username': 'nonexist', 'password': 'test'})
run_test(tc, resp.json().get('code') != 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='未校验用户是否存在', fix_suggestion='登录前查询用户是否存在', severity='中')
print_test_result(tc)

# TC-008: 错误密码
tc = TestCase(
    id='TC-008', module='认证模块', feature='用户登录', name='错误密码登录', priority=4,
    steps=[f'POST {BASE_URL}/auth/login', f'Body: {{"username":"{test_username}","password":"wrong"}}', '检查返回 code 是否不为 200'],
    expected='code!=200, 提示密码错误',
    request_detail=f'POST {BASE_URL}/auth/login'
)
resp = requests.post(f'{BASE_URL}/auth/login', json={'username': test_username, 'password': 'wrong'})
run_test(tc, resp.json().get('code') != 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='密码校验异常', fix_suggestion='检查密码比对逻辑', severity='高')
print_test_result(tc)

# TC-009: 正确登录
tc = TestCase(
    id='TC-009', module='认证模块', feature='用户登录', name='正确登录', priority=5,
    steps=[f'POST {BASE_URL}/auth/login', f'Body: {{"username":"{test_username}","password":"{test_password}"}}', '检查返回 code 是否为 200'],
    expected='code=200, 返回token',
    request_detail=f'POST {BASE_URL}/auth/login'
)
resp = requests.post(f'{BASE_URL}/auth/login', json={'username': test_username, 'password': test_password})
run_test(tc, resp.json().get('code') == 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='登录接口异常', fix_suggestion='检查登录逻辑', severity='高')
print_test_result(tc)

# TC-010: 登录暴力破解防护
tc = TestCase(
    id='TC-010', module='认证模块', feature='登录安全', name='登录频率限制', priority=5,
    steps=[f'POST {BASE_URL}/auth/login', 'Body: {"username":"admin","password":"wrong"}', '连续发送10次', '观察是否返回 429 或锁定提示'],
    expected='连续5次失败后返回 429 Too Many Requests 或提示账户锁定',
    request_detail=f'POST {BASE_URL}/auth/login (x10)'
)
blocked = False
for i in range(10):
    resp = requests.post(f'{BASE_URL}/auth/login', json={'username': 'admin', 'password': 'wrong'})
    if resp.status_code == 429 or '尝试次数' in str(resp.json().get('message', '')):
        blocked = True
        break
run_test(tc, blocked,
         actual=f'{"返回429/锁定" if blocked else "10次全部返回，无任何限制"}',
         bug_desc='无登录失败频率限制，攻击者可无限尝试暴力破解密码',
         fix_suggestion='1. 添加登录失败计数器\n2. 超过5次失败后锁定15分钟\n3. 返回429 Too Many Requests',
         severity='高')
print_test_result(tc)

# ============ 活动模块 ============
print('=' * 70)
print('活动模块测试')
print('=' * 70)

# TC-011: 未认证访问
tc = TestCase(
    id='TC-011', module='活动模块', feature='活动列表', name='未认证访问活动列表', priority=4,
    steps=['不携带 Authorization 头', f'GET {BASE_URL}/activities', '检查返回状态码'],
    expected='返回 401 或 403，拒绝访问',
    request_detail=f'GET {BASE_URL}/activities (无token)'
)
resp = requests.get(f'{BASE_URL}/activities')
run_test(tc, resp.status_code in [401, 403],
         actual=f'HTTP {resp.status_code}',
         bug_desc='未认证用户可访问受保护接口', fix_suggestion='检查 SecurityConfig 配置', severity='高')
print_test_result(tc)

# TC-012: 认证后访问
tc = TestCase(
    id='TC-012', module='活动模块', feature='活动列表', name='认证后访问活动列表', priority=5,
    steps=[f'GET {BASE_URL}/activities', f'Header: Authorization: Bearer {user_token[:30]}...', '检查返回 code'],
    expected='code=200, 返回活动列表',
    request_detail=f'GET {BASE_URL}/activities'
)
resp = requests.get(f'{BASE_URL}/activities', headers=headers)
run_test(tc, resp.json().get('code') == 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='认证后无法访问活动列表', fix_suggestion='检查 token 验证逻辑', severity='高')
print_test_result(tc)

# TC-013: 创建活动
act_data = {'title': f'测试活动{ts}', 'category': '学术讲座', 'description': '测试',
            'startTime': '2026-06-15T10:00:00', 'endTime': '2026-06-15T12:00:00',
            'location': 'A101', 'organizer': '测试', 'maxCount': 100}
tc = TestCase(
    id='TC-013', module='活动模块', feature='创建活动', name='正常创建活动', priority=5,
    steps=[f'POST {BASE_URL}/activities', f'Header: Authorization: Bearer {user_token[:30]}...', f'Body: {json.dumps(act_data, ensure_ascii=False)}', '检查返回 code'],
    expected='code=200, 返回活动详情',
    request_detail=f'POST {BASE_URL}/activities'
)
resp = requests.post(f'{BASE_URL}/activities', json=act_data, headers=headers)
r = resp.json()
activity_id = r.get('data', {}).get('id') if r.get('data') else None
run_test(tc, r.get('code') == 200,
         actual=f'code={r.get("code")}, activity_id={activity_id}',
         bug_desc='创建活动失败', fix_suggestion='检查活动创建逻辑', severity='高')
print_test_result(tc)

# TC-014: 缺少必填字段
tc = TestCase(
    id='TC-014', module='活动模块', feature='创建活动', name='缺少标题创建活动', priority=4,
    steps=[f'POST {BASE_URL}/activities', f'Header: Authorization: Bearer {user_token[:30]}...', 'Body: {"category":"test"}', '检查返回 code'],
    expected='code!=200, 提示标题不能为空',
    request_detail=f'POST {BASE_URL}/activities'
)
resp = requests.post(f'{BASE_URL}/activities', json={'category': 'test'}, headers=headers)
run_test(tc, resp.json().get('code') != 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='缺少必填字段未校验', fix_suggestion='添加 @NotBlank 注解或手动校验', severity='中')
print_test_result(tc)

# TC-015: 获取活动详情
if activity_id:
    tc = TestCase(
        id='TC-015', module='活动模块', feature='活动详情', name='获取活动详情', priority=4,
        steps=[f'GET {BASE_URL}/activities/{activity_id}', f'Header: Authorization: Bearer {user_token[:30]}...', '检查返回 code'],
        expected='code=200, 返回活动详细信息',
        request_detail=f'GET {BASE_URL}/activities/{activity_id}'
    )
    resp = requests.get(f'{BASE_URL}/activities/{activity_id}', headers=headers)
    run_test(tc, resp.json().get('code') == 200,
             actual=f'code={resp.json().get("code")}',
             bug_desc='获取活动详情失败', fix_suggestion='检查活动详情接口', severity='中')
    print_test_result(tc)

# TC-016: 审核活动（普通用户）
if activity_id:
    tc = TestCase(
        id='TC-016', module='活动模块', feature='活动审核', name='非管理员审核活动', priority=5,
        steps=[f'PUT {BASE_URL}/activities/{activity_id}/audit?action=approve', f'Header: Authorization: Bearer {user_token[:30]}... (普通用户)', '检查返回 code'],
        expected='code!=200 或 HTTP 403, 提示无权限',
        request_detail=f'PUT {BASE_URL}/activities/{activity_id}/audit?action=approve'
    )
    resp = requests.put(f'{BASE_URL}/activities/{activity_id}/audit?action=approve', headers=headers)
    run_test(tc, resp.json().get('code') != 200,
             actual=f'code={resp.json().get("code")}',
             bug_desc='任何用户都能审核活动，缺少管理员权限检查',
             fix_suggestion='1. 在 SecurityConfig 中配置审核接口需要 ADMIN 角色\n2. 在 Controller 中添加 @PreAuthorize 注解',
             severity='高')
    print_test_result(tc)

# TC-017: 无效审核操作
if activity_id:
    tc = TestCase(
        id='TC-017', module='活动模块', feature='活动审核', name='无效审核操作', priority=3,
        steps=[f'PUT {BASE_URL}/activities/{activity_id}/audit?action=invalid', f'Header: Authorization: Bearer {user_token[:30]}...', '检查返回 code'],
        expected='code!=200, 提示无效操作',
        request_detail=f'PUT {BASE_URL}/activities/{activity_id}/audit?action=invalid'
    )
    resp = requests.put(f'{BASE_URL}/activities/{activity_id}/audit?action=invalid', headers=headers)
    run_test(tc, resp.json().get('code') != 200,
             actual=f'code={resp.json().get("code")}',
             bug_desc='action 参数无验证，任意值都会执行，应只允许 approve/reject',
             fix_suggestion='在 Controller 中验证 action 值，非 approve/reject 返回 400',
             severity='中')
    print_test_result(tc)

# TC-018: 删除他人活动
if activity_id:
    # 注册另一个用户（"他人"），用其 token 去删除 user A 创建的活动
    other_ts = str(int(time.time()) + 2)[-4:]
    other_data = {'username': f'o{other_ts}', 'password': 'Test123456',
                  'confirmPassword': 'Test123456',
                  'email': f'o{other_ts}@test.com', 'phone': f'1390004{other_ts}'}
    other_resp = requests.post(f'{BASE_URL}/auth/register', json=other_data).json()
    other_token = other_resp.get('data', {}).get('token') if other_resp.get('data') else None

    tc = TestCase(
        id='TC-018', module='活动模块', feature='活动删除', name='删除他人活动', priority=5,
        steps=[
            f'用户A({test_username}) 已创建活动 id={activity_id}',
            f'注册用户B(o{other_ts}) 并获取其 token',
            f'DELETE {BASE_URL}/activities/{activity_id} (使用用户B的 token)',
            '检查返回 code',
        ],
        expected='code!=200 或 HTTP 403, 提示无权删除',
        request_detail=f'DELETE {BASE_URL}/activities/{activity_id}'
    )
    if other_token:
        other_headers = {'Authorization': f'Bearer {other_token}'}
        resp = requests.delete(f'{BASE_URL}/activities/{activity_id}', headers=other_headers)
        run_test(tc, resp.json().get('code') != 200,
                 actual=f'code={resp.json().get("code")} (用户B尝试删除用户A的活动)',
                 bug_desc='任何用户都能删除任意活动，缺少所有权检查',
                 fix_suggestion='1. 删除前验证当前用户是否为活动创建者\n2. 或验证是否为管理员',
                 severity='高')
    else:
        run_test(tc, False,
                 actual='无法创建第二个测试用户，测试环境准备失败',
                 bug_desc='无法准备测试环境（注册第二个用户失败）',
                 fix_suggestion='检查注册接口是否可用',
                 severity='高')
    print_test_result(tc)

# ============ 报名模块 ============
print('=' * 70)
print('报名模块测试')
print('=' * 70)

# TC-019: 查看我的报名
tc = TestCase(
    id='TC-019', module='报名模块', feature='我的报名', name='查看我的报名列表', priority=4,
    steps=[f'GET {BASE_URL}/registrations/me', f'Header: Authorization: Bearer {user_token[:30]}...', '检查返回 code'],
    expected='code=200, 返回当前用户的报名记录',
    request_detail=f'GET {BASE_URL}/registrations/me'
)
resp = requests.get(f'{BASE_URL}/registrations/me', headers=headers)
run_test(tc, resp.json().get('code') == 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='查看报名列表失败', fix_suggestion='检查报名列表接口', severity='中')
print_test_result(tc)

# TC-020: 报名不存在的活动
tc = TestCase(
    id='TC-020', module='报名模块', feature='活动报名', name='报名不存在的活动', priority=4,
    steps=[f'POST {BASE_URL}/registrations?activityId=99999', f'Header: Authorization: Bearer {user_token[:30]}...', '检查返回 code'],
    expected='code!=200, 提示活动不存在',
    request_detail=f'POST {BASE_URL}/registrations?activityId=99999'
)
resp = requests.post(f'{BASE_URL}/registrations?activityId=99999', headers=headers)
run_test(tc, resp.json().get('code') != 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='报名不存在活动未校验', fix_suggestion='报名前检查活动是否存在', severity='中')
print_test_result(tc)

# ============ 用户管理 ============
print('=' * 70)
print('用户管理测试')
print('=' * 70)

# TC-021: 普通用户查看用户列表
tc = TestCase(
    id='TC-021', module='用户管理', feature='用户列表', name='普通用户查看用户列表', priority=4,
    steps=[f'GET {BASE_URL}/users', f'Header: Authorization: Bearer {user_token[:30]}... (普通用户)', '检查返回 code'],
    expected='code!=200 或 HTTP 403, 提示无权限',
    request_detail=f'GET {BASE_URL}/users'
)
resp = requests.get(f'{BASE_URL}/users', headers=headers)
run_test(tc, resp.json().get('code') != 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='普通用户可以查看所有用户列表，用户信息泄露风险',
         fix_suggestion='限制 /users 接口仅管理员可访问，使用 @PreAuthorize("hasRole(\'ADMIN\')")',
         severity='高')
print_test_result(tc)

# TC-022: 更新用户状态（无效类型）
tc = TestCase(
    id='TC-022', module='用户管理', feature='用户状态', name='无效状态类型更新', priority=3,
    steps=[f'PUT {BASE_URL}/users/1/status', f'Header: Authorization: Bearer {user_token[:30]}...', 'Body: {"status": "invalid"}', '检查 HTTP 状态码'],
    expected='返回友好错误提示，非 500 服务器错误',
    request_detail=f'PUT {BASE_URL}/users/1/status'
)
resp = requests.put(f'{BASE_URL}/users/1/status', json={'status': 'invalid'}, headers=headers)
run_test(tc, resp.status_code != 500,
         actual=f'HTTP {resp.status_code}',
         bug_desc='无效状态类型导致500服务器错误',
         fix_suggestion='添加参数类型校验，返回 400 Bad Request',
         severity='中')
print_test_result(tc)

# TC-023: 获取不存在用户
tc = TestCase(
    id='TC-023', module='用户管理', feature='用户详情', name='获取不存在用户', priority=3,
    steps=[f'GET {BASE_URL}/users/99999', f'Header: Authorization: Bearer {user_token[:30]}...', '检查返回 code'],
    expected='code!=200, 提示用户不存在',
    request_detail=f'GET {BASE_URL}/users/99999'
)
resp = requests.get(f'{BASE_URL}/users/99999', headers=headers)
run_test(tc, resp.json().get('code') != 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='获取不存在用户未处理', fix_suggestion='查询结果为空时返回 404', severity='低')
print_test_result(tc)

# ============ RAG系统 ============
print('=' * 70)
print('RAG系统测试')
print('=' * 70)

# TC-024: RAG服务状态
tc = TestCase(
    id='TC-024', module='RAG系统', feature='服务状态', name='RAG服务状态检查', priority=5,
    steps=[f'GET {RAG_URL}/status/', '设置超时 5 秒', '检查 HTTP 状态码'],
    expected='HTTP 200, 返回索引状态和文档数',
    request_detail=f'GET {RAG_URL}/status/'
)
try:
    resp = requests.get(f'{RAG_URL}/status/', timeout=5)
    rag_ok = resp.status_code == 200
    run_test(tc, rag_ok,
             actual=f'HTTP {resp.status_code}',
             bug_desc='RAG服务异常', fix_suggestion='检查 RAG 服务是否启动', severity='高')
except:
    rag_ok = False
    run_test(tc, False,
             actual='连接失败，服务未启动',
             bug_desc='RAG服务未启动',
             fix_suggestion='启动 RAG 服务: cd campus-rag && python main.py',
             severity='高')
print_test_result(tc)

if rag_ok:
    # TC-025: 空问题
    tc = TestCase(
        id='TC-025', module='RAG系统', feature='对话功能', name='空问题发送', priority=4,
        steps=[f'POST {RAG_URL}/chat/', 'Body: {"question": ""}', '检查 HTTP 状态码'],
        expected='HTTP 400, 提示请输入问题',
        request_detail=f'POST {RAG_URL}/chat/'
    )
    resp = requests.post(f'{RAG_URL}/chat/', json={'question': ''})
    run_test(tc, resp.status_code == 400,
             actual=f'HTTP {resp.status_code}',
             bug_desc='空问题未校验', fix_suggestion='添加空字符串校验', severity='低')
    print_test_result(tc)

    # TC-026: 创建对话
    tc = TestCase(
        id='TC-026', module='RAG系统', feature='对话管理', name='创建新对话', priority=4,
        steps=[f'POST {RAG_URL}/conversations/', 'Body: {"title": "测试"}', '检查 HTTP 状态码'],
        expected='HTTP 201, 包含对话ID',
        request_detail=f'POST {RAG_URL}/conversations/'
    )
    resp = requests.post(f'{RAG_URL}/conversations/', json={'title': '测试'})
    run_test(tc, resp.status_code == 201,
             actual=f'HTTP {resp.status_code}',
             bug_desc='创建对话失败', fix_suggestion='检查对话创建接口', severity='中')
    print_test_result(tc)

    # TC-027: 删除不存在文档
    tc = TestCase(
        id='TC-027', module='RAG系统', feature='文档管理', name='删除不存在文档', priority=3,
        steps=[f'POST {RAG_URL}/delete/', 'Body: {"filename": "nonexist.txt"}', '检查 HTTP 状态码'],
        expected='HTTP 400, 提示文件不存在',
        request_detail=f'POST {RAG_URL}/delete/'
    )
    resp = requests.post(f'{RAG_URL}/delete/', json={'filename': 'nonexist.txt'})
    run_test(tc, resp.status_code == 400,
             actual=f'HTTP {resp.status_code}',
             bug_desc='删除不存在文档未处理', fix_suggestion='检查文件是否存在再删除', severity='低')
    print_test_result(tc)

# TC-028: chunk_text无限循环风险
tc = TestCase(
    id='TC-028', module='RAG系统', feature='文本分块', name='chunk_size等于overlap死循环', priority=3,
    steps=[f'POST {RAG_URL}/upload_text/', 'Body: {"text":"测试文本x20", "strategy":"fixed", "chunk_size":100, "overlap":100}', '设置超时 3 秒', '观察是否超时'],
    expected='正常返回或提示参数错误',
    request_detail=f'POST {RAG_URL}/upload_text/'
)
if rag_ok:
    try:
        resp28 = requests.post(f'{RAG_URL}/upload_text/', json={
            'text': '这是一段测试文本。' * 20,
            'filename': 'test.txt',
            'strategy': 'fixed',
            'chunk_size': 100,
            'overlap': 100
        }, timeout=3)
        run_test(tc, resp28.status_code == 200,
                 actual=f'HTTP {resp28.status_code}',
                 bug_desc='chunk_size=overlap 时 fixed 策略无限循环',
                 fix_suggestion='添加参数验证: overlap 必须小于 chunk_size',
                 severity='中')
    except requests.exceptions.Timeout:
        run_test(tc, False,
                 actual='请求超时（3秒），函数进入死循环',
                 bug_desc='chunk_size=overlap 时 fixed 策略无限循环',
                 fix_suggestion='添加参数验证: overlap 必须小于 chunk_size',
                 severity='中')
    except Exception as e:
        run_test(tc, False,
                 actual=f'请求异常: {e}',
                 bug_desc='RAG服务请求异常',
                 fix_suggestion='检查 RAG 服务状态',
                 severity='中')
else:
    run_test(tc, False,
             actual='RAG服务未启动，无法测试',
             bug_desc='RAG服务未启动，无法测试',
             fix_suggestion='启动 RAG 服务后重新测试',
             severity='中')
print_test_result(tc)

# ============ 安全测试 ============
print('=' * 70)
print('安全测试')
print('=' * 70)

# TC-029: SQL注入
tc = TestCase(
    id='TC-029', module='安全测试', feature='SQL注入', name='SQL注入防护', priority=5,
    steps=[f'POST {BASE_URL}/auth/login', 'Body: {"username":"admin\' OR \'1\'=\'1","password":"x"}', '检查返回 code'],
    expected='code!=200, SQL注入无效',
    request_detail=f'POST {BASE_URL}/auth/login'
)
resp = requests.post(f'{BASE_URL}/auth/login', json={'username': "admin' OR '1'='1", 'password': 'x'})
run_test(tc, resp.json().get('code') != 200,
         actual=f'code={resp.json().get("code")}',
         bug_desc='存在SQL注入漏洞', fix_suggestion='使用参数化查询（MyBatis #{}）', severity='严重')
print_test_result(tc)

# TC-030: JWT篡改
tc = TestCase(
    id='TC-030', module='安全测试', feature='JWT安全', name='JWT token篡改', priority=5,
    steps=[f'GET {BASE_URL}/activities', 'Header: Authorization: Bearer fake.token.here', '检查 HTTP 状态码'],
    expected='返回 401 或 403，拒绝伪造 token',
    request_detail=f'GET {BASE_URL}/activities'
)
resp = requests.get(f'{BASE_URL}/activities', headers={'Authorization': 'Bearer fake.token.here'})
run_test(tc, resp.status_code in [401, 403],
         actual=f'HTTP {resp.status_code}',
         bug_desc='JWT 篡改未检测', fix_suggestion='检查 JWT 签名验证逻辑', severity='严重')
print_test_result(tc)

# TC-031: 禁用用户后token有效性
tc = TestCase(
    id='TC-031', module='安全测试', feature='JWT安全', name='禁用用户后token失效', priority=5,
    steps=['注册新用户 dis_test', '获取 token', f'GET {BASE_URL}/activities (确认 token 有效)', '管理员禁用该用户 (status=0)', f'GET {BASE_URL}/activities (再次使用原 token)', '检查是否返回 401'],
    expected='禁用后 token 立即失效，返回 401',
    request_detail=f'PUT {BASE_URL}/users/{{id}}/status + GET {BASE_URL}/activities'
)
if user_token:
    disable_ts = str(int(time.time()))[-4:]
    disable_data = {'username': f'dis{disable_ts}', 'password': 'Test123456', 'confirmPassword': 'Test123456',
                    'email': f'dis{disable_ts}@test.com', 'phone': f'1380000{disable_ts}'}
    resp_d = requests.post(f'{BASE_URL}/auth/register', json=disable_data)
    rd = resp_d.json()
    if rd.get('code') == 200 and rd.get('data'):
        disable_token = rd['data']['token']
        disable_uid = rd['data']['id']
        disable_headers = {'Authorization': f'Bearer {disable_token}'}
        resp_check = requests.get(f'{BASE_URL}/activities', headers=disable_headers)
        token_works_before = resp_check.json().get('code') == 200
        requests.put(f'{BASE_URL}/users/{disable_uid}/status', json={'status': 0}, headers=headers)
        resp_after = requests.get(f'{BASE_URL}/activities', headers=disable_headers)
        token_rejected = resp_after.status_code in [401, 403]
        run_test(tc, token_works_before and token_rejected,
                 actual=f'禁用前token有效={token_works_before}, 禁用后token被拒={token_rejected}',
                 bug_desc='禁用用户后JWT token仍然有效，缺少token失效机制',
                 fix_suggestion='在 JwtAuthenticationFilter 中检查用户 status，status!=1 时拒绝',
                 severity='严重')
        requests.put(f'{BASE_URL}/users/{disable_uid}/status', json={'status': 1}, headers=headers)
    else:
        run_test(tc, False, actual='无法创建测试用户', bug_desc='无法创建测试用户', fix_suggestion='检查注册接口', severity='严重')
else:
    run_test(tc, False, actual='无可用token', bug_desc='无可用token', fix_suggestion='检查注册接口', severity='严重')
print_test_result(tc)

# TC-032: XSS防护
tc = TestCase(
    id='TC-032', module='安全测试', feature='XSS防护', name='活动标题XSS注入', priority=4,
    steps=[f'POST {BASE_URL}/activities', f'Header: Authorization: Bearer {user_token[:30]}...', 'Body: {"title":"<script>alert(1)</script>", ...}', '检查返回的 title 是否被转义'],
    expected='标题应被转义为 &lt;script&gt;',
    request_detail=f'POST {BASE_URL}/activities'
)
if user_token:
    xss_data = {'title': '<script>alert(1)</script>', 'category': 'test', 'description': 'xss',
                'startTime': '2026-06-15T10:00:00', 'endTime': '2026-06-15T12:00:00',
                'location': 'test', 'organizer': 'test'}
    resp = requests.post(f'{BASE_URL}/activities', json=xss_data, headers=headers)
    title = resp.json().get('data', {}).get('title', '') if resp.json().get('data') else ''
    run_test(tc, '<script>' not in title,
             actual=f'返回标题: {title}',
             bug_desc='活动标题未转义HTML标签，存在XSS攻击风险',
             fix_suggestion='1. 入库前对用户输入进行HTML转义\n2. 前端使用 v-text 而非 v-html',
             severity='高')
else:
    run_test(tc, False, actual='无可用token', bug_desc='无可用token', fix_suggestion='检查注册接口', severity='高')
print_test_result(tc)

# ============ 汇总 ============
passed = sum(1 for tc in all_tests if tc.result == 'Pass')
failed = sum(1 for tc in all_tests if tc.result == 'Fail')
total = len(all_tests)

print()
print('=' * 70)
print('                        测试结果汇总')
print('=' * 70)
print(f'测试环境: {BASE_URL}')
print(f'测试账号: {test_username} / {test_password}')
print(f'通过: {passed}')
print(f'失败: {failed}')
print(f'总计: {total}')
print(f'通过率: {passed / total * 100:.1f}%' if total > 0 else '')
print()

# 输出 bug 报告
print_bug_report()

print('=' * 70)

# ============ 自动生成 Excel 测试报告 ============
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- Sheet 1: 通过的测试用例 ---
    ws_pass = wb.active
    ws_pass.title = "通过的测试用例"
    pass_headers = ['编号', '模块', '功能名称', '测试用例名称', '级别', '测试环境', '测试账号', '测试步骤', '请求详情', '预期结果', '实际结果', '测试结果']
    for col, h in enumerate(pass_headers, 1):
        cell = ws_pass.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    row_idx = 2
    for tc in all_tests:
        if tc.result == 'Pass':
            values = [tc.id, tc.module, tc.feature, tc.name, tc.priority,
                      BASE_URL, f'{test_username} / {test_password}',
                      '\n'.join(tc.steps), tc.request_detail,
                      tc.expected, tc.actual, 'Pass']
            for col_idx, val in enumerate(values, 1):
                cell = ws_pass.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if col_idx == len(values):
                    cell.fill = pass_fill
                    cell.font = Font(bold=True, color="006100")
            row_idx += 1

    for i, w in enumerate([12, 10, 12, 20, 6, 25, 20, 40, 35, 25, 25, 8], 1):
        ws_pass.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 2: 失败的测试用例(BUG) ---
    ws_fail = wb.create_sheet("失败的测试用例(BUG)")
    fail_headers = ['编号', '模块', '功能名称', '测试用例名称', '级别', '测试环境', '测试账号', '测试步骤', '请求详情', '预期结果', '实际结果', '测试结果', 'BUG描述', '修复建议', '严重程度']
    for col, h in enumerate(fail_headers, 1):
        cell = ws_fail.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    row_idx = 2
    for tc in all_tests:
        if tc.result == 'Fail':
            values = [tc.id, tc.module, tc.feature, tc.name, tc.priority,
                      BASE_URL, f'{test_username} / {test_password}',
                      '\n'.join(tc.steps), tc.request_detail,
                      tc.expected, tc.actual, 'Fail',
                      tc.bug_desc, tc.fix_suggestion, tc.severity]
            for col_idx, val in enumerate(values, 1):
                cell = ws_fail.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if col_idx == 12:
                    cell.fill = fail_fill
                    cell.font = Font(bold=True, color="9C0006")
            row_idx += 1

    for i, w in enumerate([12, 10, 12, 20, 6, 25, 20, 40, 35, 25, 25, 8, 35, 35, 10], 1):
        ws_fail.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 3: 测试汇总 ---
    ws_summary = wb.create_sheet("测试汇总")
    summary_headers = ['统计项', '数量', '占比']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    summary_data = [
        ['测试用例总数', total, '100%'],
        ['通过用例数', passed, f'{passed/total*100:.1f}%' if total else '0%'],
        ['失败用例数(BUG)', failed, f'{failed/total*100:.1f}%' if total else '0%'],
        ['', '', ''],
        ['测试环境', BASE_URL, ''],
        ['RAG服务', RAG_URL, ''],
        ['测试账号', f'{test_username} / {test_password}', ''],
        ['测试时间', TEST_TIME, ''],
    ]
    for r, data in enumerate(summary_data, 2):
        for c, val in enumerate(data, 1):
            cell = ws_summary.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for i, w in enumerate([20, 30, 15], 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = w

    # 保存文件（带日期时间戳，存放到 test 目录）
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_file = os.path.join(output_dir, f'测试报告_{timestamp}.xlsx')
    wb.save(output_file)
    print(f'Excel 报告已生成: {output_file}')
except ImportError:
    print('提示: 安装 openpyxl 可自动生成 Excel 报告 (pip install openpyxl)')
except Exception as e:
    print(f'生成 Excel 失败: {e}')
